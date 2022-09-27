'''
文件名：GetAllPaperText-Plus.py
这个是 02、GetAllPaperText.py 的升级版，目前是 完成于 20220915
针对对象是经 GetJourInfo.py 和 AddJCR2JourInfo.py 生成的 JournalInfo.xlsx，
此表格含有 CCF对应的所有期刊和会议，且期刊经中科院对应接口获取其分区信息
目前有四张子表，两张原数据，两张整理后
'''
from openpyxl import load_workbook

wb = load_workbook('中国计算机学会推荐国际学术会议和期刊目录-2019/JournalInfo.xlsx')
sheets = wb.worksheets
# 这里就是进行子表的复制
if '区块链对应期刊' in wb.sheetnames:
    print('区块链对应期刊 已存在')
    NewSheets1 = wb['区块链对应期刊']
else:
    NewSheets1 = wb.copy_worksheet(wb['期刊信息'])
    NewSheets1.title = '区块链对应期刊'
if '区块链对应会议' in wb.sheetnames:
    print('区块链对应会议 已存在')
    NewSheets2 = wb['区块链对应会议']
else:
    NewSheets2 = wb.copy_worksheet(wb['会议信息'])
    NewSheets2.title = '区块链对应会议'
wb.save('中国计算机学会推荐国际学术会议和期刊目录-2019/JournalInfo.xlsx')

# 获取表格已使用列数
def GetTrueLen(list):
    theLen = len(list)
    while (theLen):
        if list[theLen - 1] != None:
            return theLen
        else:
            theLen = theLen - 1

# 获取期刊对应子网页链接
import requests
import re

# 这里就是用re正则表达中最麻烦的部分了，因为不同的期刊格式不一样，一个正则表达式可能只对某一期刊有效
# 换言之就是其他期刊在套同一正则表达式不一定拿到数据
# 所以这里用了三个正则表达式
patrenForGetUrl1 = '<li><a href=\"(.*?)\">.*?Volume(.*?)</a></li>'
patrenForGetUrl2 = '<a href=\"(.*?)\">(.*?)</a>'
patrenForGetUrl3 = '<a href=\"(.*?)\">(.*?)Volume(.*?)</a>'

# 这里就是爬主网页数据
def GetMainPage(WebUrl, patrenForGetUrl):
    headers = {
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/105.0.0.0 Safari/537.36 Edg/105.0.1343.27',
        'Connection': 'close'
    }
    response = requests.get(url=WebUrl, headers=headers)
    response.encoding = 'utf-8'
    html_str = response.text
    # print(html_str)
    Get_Page = re.compile(patrenForGetUrl).findall(html_str)
    print(Get_Page)
    response.close()
    return Get_Page


# 用于将获取到的列表进行归一化，参数包括：爬虫获取到的一堆url，自身主页对应的url，用来进行匹配
# 这里是针对patrenForGetUrl2 获取到的list进行处理
# 格式为：（url,year,num）,如：('https://dblp.uni-trier.de/db/journals/ivc/ivc117.html', '2022,117')
# 这里格式可改，在NewVal 那一行自行修改
def StandForm2(list, url):
    SetYear = 2022 #这个是因为存在一些期刊没有年份只有卷数，我默认为2022，当卷数开始降低就说明年份开始自减
    for inx, val in enumerate(list):
        if re.findall(url, val[0], re.I):
            if re.split('-', list[inx + 1][1])[-1].isdigit():  # 用于对网页中存在部分格式为xx-xx的解析
                if inx + 1 != len(list):
                    NewVal = (val[0], str(SetYear) + ',' + str(val[1]),)
                    if int(re.split('-', list[inx][1])[0]) > int(re.split('-', list[inx + 1][1])[0]):
                        SetYear = SetYear - 1
                MainPape.append(NewVal)
            else:
                NewVal = (val[0], str(SetYear) + ',' + str(val[1]),)
                MainPape.append(NewVal)
                break
    print(MainPape)
    return MainPape


# 这里是针对patrenForGetUrl1解析出来的数据进行归一化处理
def StandForm1(list):
    for inx, val in enumerate(list):
        temp = re.sub(' ', '', val[1])
        # print(temp)
        if re.findall(',', temp, re.I):
            temp2 = re.split(',', temp)
            UrlInfo = temp2[1] + ',' + temp2[0]
        elif re.findall(':', temp, re.I):
            temp2 = re.split(':', temp)
            UrlInfo = temp2[1] + ',' + temp2[0]
        else:
            UrlInfo = temp
        NewVal = (val[0], UrlInfo,)
        MainPape.append(NewVal)
    print(MainPape)
    return MainPape


# 这里是针对patrenForGetUrl3解析出来的数据进行归一化处理
def StandForm3(list):
    for inx, val in enumerate(list):
        year = re.sub(' ', '', val[1])
        num = re.sub(" ", '', val[2])
        year = re.sub(':', '', year)
        # print(year,num)
        UrlInfo = year + ',' + num
        NewVal = (val[0], UrlInfo,)
        MainPape.append(NewVal)
    print(MainPape)
    return MainPape


'''
通过传入关键字进行匹配
# 这里可以做很多有意思的操作，比如我这里是只匹配一个KW
# 通过在这里修改判断，可以做多重关键字的判断
# 比如如果KW1没有匹配到，那能不能匹配kw2 再返回，或者
# KW1匹配到了，再匹配KW2，都匹配到了再返回
'''
def BlockFind(kw, info):
    if re.findall(kw, info, re.I):
        return info


# 这里是期刊子页面的论文搜索
def GetThePaperTitle(MainPape, KW):
    # import io
    # import sys
    # import urllib.request
    # sys.stdout = io.TextIOWrapper(sys.stdout.buffer,encoding='utf8') #改变标准输出的默认编码
    countofPaper = 0;
    countOFKW = 0;
    paperInfo = []
    headers = {
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/105.0.0.0 Safari/537.36 Edg/105.0.1343.27',
        'Connection': 'close'
    }
    from bs4 import BeautifulSoup
    ForYearwithKW = 0
    for i in MainPape:
        nextUrl = i[0]
        nextUrlText = i[1]
        print(nextUrl, nextUrlText)
        year = re.findall('\d+', re.split(',', nextUrlText)[0])[-1]
        # 这里是为了记录每一次爬取的论文信息
        # with open('txt-src/AllPaperInfo.txt', 'a',encoding='utf-8') as file:
        #     file.write('\n'+'## '+year+'\n')
        #     file.close()
        # print(year)
        PaperInfo = requests.get(url=nextUrl, headers=headers)  # GET请求
        PaperInfo.encoding = ('utf-8')
        soup = BeautifulSoup(PaperInfo.text, 'lxml')
        temp1 = soup.find_all('li', itemtype="http://schema.org/ScholarlyArticle")
        ForYearNum = 0
        Title = []
        for i in temp1:
            u = i.find('div', class_="head").find('a')
            if u == None:
                continue
            # print(u['href'])
            t = i.find('span', class_="title", itemprop="name")
            ForYearNum = ForYearNum + 1
            Title.append(t.text + '\n')
            if BlockFind(KW, t.text):
                # 匹配成功,写入表格中
                countOFKW = countOFKW + 1
                NewVal = (year, u['href'], t.text,)
                # print(type(year),type(u['href']),type(t.text))
                paperInfo.append(NewVal)
            countofPaper = countofPaper + 1
        # 这里是将每一次爬取到的论文标题按顺序写入，在写入前先写入 当前论文的数目
        # with open('txt-src/AllPaperInfo.txt', 'a', encoding='utf-8') as file:
        #     file.write('\n' + '## ' + year + ',' + str(ForYearNum) + '\n')
        #     for i in Title:
        #         file.write(i)
        #     file.close()
        print("第 " + year + " 年对应论文数：" + str(ForYearNum))
    print("总论文数：" + str(countofPaper) + ", 对应关键字有效匹配数：" + str(countOFKW))
    paperInfo.append(str(countofPaper))  # 倒数第二位是所有论文的数字
    paperInfo.append(str(countOFKW))  # 最后一位是KW对应的论文数字
    return paperInfo

## 注意这里是因为网络波动，会存在中断情况，这时候如果你没有设置GetThePaperTitle写入txt
## 只需要把下面的三个参数修改成最后一次打印输出的结果即可
# 开始期刊
# 比如第一次失败：
# 修改AllPaperNum = 378916 ；AllPaperwithKW =657 ；TruePaperwithKW =190695
'''
当前已遍历论文总数为： 378916 ,已遍历对应关键字的论文总数为： 657 ,拥有KW的期刊论文总数为： 190695
9 ['人工智能', 'A-9.088', 'AI', 'Artificial Intelligence', 'Elsevier', 'https://dblp.uni-trier.de/db/journals/ai/', '0004-3702', '计算机：人工智能', 2, None, None, None, None, None, None, None, None]
计算机：人工智能

'''
# 第二次失败：
'''
当前已遍历论文总数为： 446906 ,已遍历对应关键字的论文总数为： 674 ,拥有KW的期刊论文总数为： 199096
13 ['人工智能', 'C-5.795', 'DSS', 'Decision Support Systems', 'Elsevier',
'''
# 最终：当前已遍历论文总数为： 622369 ,已遍历对应关键字的论文总数为： 808 ,拥有KW的期刊论文总数为： 273558


AllPaperNum = 0  # 这是记录所有论文的个数
AllPaperwithKW = 0  # 这是记录对应KW的个数
TruePaperwithKW = 0  # 这是记录存有KW下的总论文数
KW = "blockchain" ## 这里的关键字是可以修改的
SetRows = 0
rows = NewSheets1.rows
columns = NewSheets1.columns
for row in rows:
    MainPape = []
    row_val = [col.value for col in row]
    # 对应行的真实列数
    GetCol = GetTrueLen(row_val)
    print(GetCol, row_val)
    print(row_val[GetCol - 2])  # 这里输出当前行的最后一位，判断是否已经写入，正常是字符串，然后写完后就是所有论文的数目
    SetRows = SetRows + 1
    if row_val[GetCol - 2].isdigit() != True:
    	# 这里是写入 期刊的信息，格式有点像 makedown
        # with open('txt-src/AllPaperInfo.txt', 'a', encoding='utf-8') as file:
        #     file.write('\n' + '# ')
        #     for i in range(GetCol):
        #         if type(row_val[i]) != str:
        #             file.write(str(row_val[i]) + '\t')
        #         else:
        #             file.write(row_val[i] + '\t')
        url = row_val[5]
        # print(url)
        # url = 'https://dblp.uni-trier.de/db/journals/jpdc/'
        MainPape = StandForm1(GetMainPage(url, patrenForGetUrl1))
        if MainPape == []:
            GetPage = GetMainPage(url, patrenForGetUrl2)
            MainPape = StandForm2(GetPage, url)
            if MainPape == []:
                MainPape = StandForm3(GetMainPage(url, patrenForGetUrl3))
        # 此处对获取到的MianPape做子页获取
    else:
        continue

    paperInfo = GetThePaperTitle(MainPape, KW) # 这里获取到期刊对应的所有论文
    NowPaperNum = int(paperInfo[-2])  # 获取所有论文的总数
    NowPaperwithKW = int(paperInfo[-1])  # 获取对应KW的论文数
    if NowPaperwithKW != 0:
        TruePaperwithKW = TruePaperwithKW + NowPaperNum  # 计算涉及到关键字的期刊的论文总数
        for i in range(NowPaperwithKW):
        	# 这里是按年、论文标题，论文标题对应格子里设置超链接
            print("第 " + str(SetRows) + " 行，第 " + str(GetCol + 1 + i * 2) + " 列 在写入")
            NewSheets1.cell(row=SetRows, column=GetCol + 1 + i * 2).value = paperInfo[i][0]
            NewSheets1.cell(row=SetRows, column=GetCol + 2 + i * 2).value = paperInfo[i][2]
            NewSheets1.cell(row=SetRows, column=GetCol + 2 + i * 2).hyperlink = paperInfo[i][1]
    AllPaperNum = AllPaperNum + NowPaperNum  # 计算当前所有期刊的论文总数
    AllPaperwithKW = AllPaperwithKW + NowPaperwithKW  # 计算当前对应关键字的所有期刊的论文总数
    NewSheets1.cell(row=SetRows, column=GetCol + NowPaperwithKW * 2 + 1).value = paperInfo[-2]  # 写入当前期刊的论文总数
    NewSheets1.cell(row=SetRows, column=GetCol + 2 + NowPaperwithKW * 2).value = paperInfo[-1]  # 写入对应kw的论文总数
    wb.save('中国计算机学会推荐国际学术会议和期刊目录-2019/JournalInfo.xlsx')
    print("当前已遍历论文总数为： " + str(AllPaperNum) + ' ,已遍历对应关键字的论文总数为： ' + str(AllPaperwithKW) + " ,拥有KW的期刊论文总数为： " + str(
        TruePaperwithKW))
    # 这里是把每次的运行结果进行记录
    # with open('txt-src/logs.txt', 'a', encoding='utf-8') as f:
    #         # f.write("目前已遍历到期刊：")
    #     f.write('\n' + '# ')
    #     for i in range(GetCol):
    #         if type(row_val[i]) != str:
    #             f.write(str(row_val[i]) + '\t')
    #         else:
    #             f.write(row_val[i] + '\t')
    #     f.write('\n' + "## 当前已遍历论文总数为： ")
    #     f.write(str(AllPaperNum))
    #     f.write(' ,已遍历对应关键字的论文总数为： ')
    #     f.write(str(AllPaperwithKW))
    #     f.write(" ,拥有KW的期刊论文总数为： ")
    #     f.write(str(TruePaperwithKW) + '\n')

