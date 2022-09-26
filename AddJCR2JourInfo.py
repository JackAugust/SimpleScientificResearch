# 参考 https://blog.csdn.net/weixin_41819529/article/details/88706688
# AddJCR2JourInfo.py
import requests
import json

from openpyxl import load_workbook
wb = load_workbook('中国计算机学会推荐国际学术会议和期刊目录-2019/JournalInfo.xlsx')
sheets = wb.worksheets
Journal = wb['期刊信息']
url = 'http://webapi.fenqubiao.com/api/journal'

# 这里是在子表前插入一列用来添加大类信息
Journal.insert_cols(1,1)
JourClass = ['计算机体系结构/并行与分布计算/存储系统','计算机网络','网络与信息安全','软件工程/系统软件/程序设计语言','数据库/数据挖掘/内容检索','计算机科学理论','计算机图形学与多媒体','人工智能','人机交互与普适计算','交叉/综合/新兴']
# 通过保存此时表已完成在前插入一列
wb.save('中国计算机学会推荐国际学术会议和期刊目录-2019/JournalInfo.xlsx')
print(Journal.max_row,Journal.max_column)
rows = Journal.rows
columns = Journal.columns
countA = 0
countB = 0
countC = 0

ClassNum = 0
SetRows = 0 # 行数

for row in rows:
    # print(row)
    row_val = [col.value for col in row]
    issn = row_val[6]
    print(issn)
    ## 这里往下参考的博客，其实就是生成一定格式的url来获取数据
    search = 'year=2021&abbr='+issn;
    req = requests.get(url='%s%s%s' % (url, '?', search))
    req.encoding='utf-8'
    result = json.loads(req.text)
    print(result)
    SetRows = SetRows + 1
    # 这里是开始在第一列添加大类信息
    Journal.cell(row=SetRows, column=1).value = JourClass[ClassNum]
    strValue = Journal.cell(row=SetRows,column=2).value
    # 这里的三个判断就是为了算当前有多少个A\B\C类
    if strValue == 'A':
        countA = countA + 1;
    elif strValue == 'B':
        countB = countB + 1
    elif strValue == 'C':
        countC = countC + 1;
    if SetRows<Journal.max_row:
    # 这里是当此行是C类，下一行是A类，说明要换大类，就是这个意思
        if strValue=="C" and Journal.cell(row=SetRows+1,column=2).value=="A":
            ClassNum = ClassNum + 1
    if len(result) == 1:
        continue
    else:
        print("期刊名称：",result['Title'])
        Journal.cell(row=SetRows,column=2).value=Journal.cell(row=SetRows,column=2).value+'-'+str(result['Indicator']['ImpactFactor'])
        # 这里是把大类和分区写到每行的后面
        for index,jcr in enumerate(result['JCR']):
            print (index,"类别：%s,分区：%s"%(jcr["NameCN"],jcr["Section"]))
            Journal.cell(row=SetRows,column=index*2+8).value=jcr["NameCN"]
            Journal.cell(row=SetRows, column=index * 2 + 9).value = jcr["Section"]
        wb.save('中国计算机学会推荐国际学术会议和期刊目录-2019/JournalInfo.xlsx')
print('A类：' + str(countA) + ' B类： ' + str(countB) + " C类：" + str(countC))
# A类：32 B类： 102 C类：105
# 这里是会议。没有分区需要添加
Meet = wb['会议信息']
SetRows = 0
ClassNum = 0
countA = 0
countB = 0
countC = 0
Meet.insert_cols(1,1)
wb.save('中国计算机学会推荐国际学术会议和期刊目录-2019/JournalInfo.xlsx')
rows = Meet.rows
columns = Meet.columns
for row in rows:
    row_val = [col.value for col in row]
    SetRows = SetRows + 1
    strValue = Meet.cell(row=SetRows, column=2).value
    if strValue == 'A':
        countA = countA + 1;
    elif strValue == 'B':
        countB = countB + 1
    elif strValue == 'C':
        countC = countC + 1;
    Meet.cell(row=SetRows, column=1).value = JourClass[ClassNum]
    if SetRows < Meet.max_row:
        if strValue == "C" and Meet.cell(row=SetRows + 1, column=2).value == "A":
            ClassNum = ClassNum + 1
wb.save('中国计算机学会推荐国际学术会议和期刊目录-2019/JournalInfo.xlsx')
print('A类：' + str(countA) + ' B类： ' + str(countB) + " C类：" + str(countC))
# A类：51 B类： 121 C类：156