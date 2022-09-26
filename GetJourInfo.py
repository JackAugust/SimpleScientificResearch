# 此时仍采用简单的requests 和 re
# GetJourInfo.py
import requests
import re
'''
# 该py文件用于将期刊信息进行规格化，通过整理好的txt文件来获取期刊信息和类比，并添加
# 期刊对应的ISSN号用于期刊分区
'''
# 获取期刊对应的ISSN
## 这里是爬虫的部分，注意这里我是爬取对应网页的ISSN部分。不同的是我写成了函数：GetJournISSN()
## 返回的是爬取的ISSN和一些其他数据
headers = {
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/105.0.0.0 Safari/537.36 Edg/105.0.1343.27',
    'Connection': 'close'
}
def GetJournISSN(url):
    getUrl = requests.get(url=url, headers=headers,timeout=10)
    getUrl.encoding='utf-8'
    JournalInfopatren = '<div class="hide-body"><ul><li><em>.*?</em> <a href=".*?">(.*?)</a>'
    getUrl.close()
    return re.compile(JournalInfopatren, re.S).findall(getUrl.text)

'''
文件处理
'''
from openpyxl import load_workbook
with open('中国计算机学会推荐国际学术会议和期刊目录-2019/03、处理后-中国计算机学会推荐国际学术会议和期刊目录-2019-去引号.txt','r',encoding='utf-8') as file:
    getJournal = file.readlines();

# 分表说明
## 这里是为了把数据按期刊和会议的方式分开存表
wb = load_workbook('中国计算机学会推荐国际学术会议和期刊目录-2019/JournalInfo.xlsx')
sheets = wb.worksheets
Journal = sheets[0]
Meet = sheets[1]
if '期刊信息' in wb.sheetnames:
    print('期刊信息 已存在')
    NewSheets1 = wb['期刊信息']
else:
    NewSheets1 = wb.create_sheet('期刊信息')

if '会议信息' in wb.sheetnames:
    print('会议信息 已存在')
    NewSheets2 = wb['会议信息']
else:
    NewSheets2 = wb.create_sheet('会议信息')

## 这里是计数，确定一共有多少ABC
countA = 0
countB = 0
countC = 0
# ֱ直接设置类别说明
Data = [] # 这里的Data存的就只有期刊或会议的数据了，完成清洗
for i in range(len(getJournal)):
    if re.search('A 类', getJournal[i]):
        level = 'A'
    elif re.search('B 类', getJournal[i]):
        level = 'B'
    elif re.search('C 类', getJournal[i]):
        level = 'C'
    NewData = re.split('\t',re.sub('\n','',getJournal[i]))
    if len(NewData) == 5:
        NewData[0] = level
        if level== 'A':
            countA = countA + 1
        elif level == 'B':
            countB = countB + 1
        elif level == 'C':
            countC = countC + 1
        Data.append(NewData)
print('A类：'+str(countA) + ' B类： '+str(countB)+" C类："+str(countC))

JourCount = 0 # 期刊类别计数
MeetCount = 0 # 会议类别计数

# 主函数，将Data数据写入表格
SetRows1 = 1; # 期刊子表的行数
SetRows2 = 1; # 会议子表的行数
for i in Data:
    col = 1;
    print(i)
    if re.search('http://dblp.uni-trier.de/db/journals/',i[4]):
        # 这里开始把数据写入子表 NewSheets1 里
        for j in range(len(i)):
            NewSheets1.cell(row=SetRows1,column=col+j).value=i[j]
        issn = GetJournISSN(i[4]) # 这里就通过爬虫获取到期刊对应的ISSN
        if issn:
            print(issn)
            NewSheets1.cell(row=SetRows1,column=col+5).value=issn[0]
        SetRows1 = SetRows1 + 1
        JourCount = JourCount + 1
        # print(info1)
    elif re.search('http://dblp.uni-trier.de/db/conf/',i[4]):
        # 这里开始把数据写入子表 NewSheets2 里
        for j in range(len(i)):
            NewSheets2.cell(row=SetRows2, column=col + j).value = i[j]
        SetRows2 = SetRows2 + 1
        MeetCount = MeetCount + 1
    wb.save('中国计算机学会推荐国际学术会议和期刊目录-2019/JournalInfo.xlsx')
print('期刊数'+str(JourCount)+' ,会议数 '+str(MeetCount))
print('A类：'+str(countA) + ' B类： '+str(countB)+" C类："+str(countC))

# 总数：
## 期刊总数： 266 , 32, 111, 123
## 会议总数： 347 , 53, 125, 169
# 经清洗后：
    # 期刊数238 ,会议数 328
    # A类：96 B类： 245 C类：297
# 运行结果：
'''
['A', 'TOCS', 'ACM Transactions on Computer Systems', 'ACM', 'http://dblp.uni-trier.de/db/journals/tocs/']
['0734-2071']
['A', 'TOS', 'ACM Transactions on Storage', 'ACM', 'http://dblp.uni-trier.de/db/journals/tos/']
['1553-3077']
'''