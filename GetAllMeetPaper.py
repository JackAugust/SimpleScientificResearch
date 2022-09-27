'''
该文件用于获取会议论文，前提：
GetJourlInfo 写入 “会议信息”
直接在此文件中进行爬取到判断到写入即可，因网络中断可直接运行，建议把 日志文件：
txt-src/AllMeetInfo.txt 内的重复数据删除 即可
'''

from openpyxl import load_workbook
import re
import requests
from bs4 import BeautifulSoup

# 获取表格已使用列数
def GetTrueLen(list):
    theLen = len(list)
    while (theLen):
        if list[theLen - 1] != None:
            return theLen
        else:
            theLen = theLen - 1
# 获取网页数据
def GetUrlInfo(url):
    headers = {
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/105.0.0.0 Safari/537.36 Edg/105.0.1343.27',
        'Connection': 'close'
    }
    response = requests.get(url=url, headers=headers)
    response.encoding = 'utf-8'
    return response

# 获取会议的主页链接，输入参数为url和设置年份
# 本来思路是 如果会议的时间小于区块链提出时间就没必要遍历了，不然数据量太大
def GetMeetMainInfo(url,SetYear):
    response = GetUrlInfo(url)
    soup = BeautifulSoup(response.text, 'lxml')
    temp1 = soup.find_all('li', itemtype="http://schema.org/Book")
    MeetPaper=[]
    oldyaer = SetYear
    for i in temp1:
        u = i.find('div', class_="head").find('a')
        if u == None:
            continue
        t = i.find('span', class_="title", itemprop="name")
        getYear = re.findall("\d+", t.text)
        if len(getYear) != 0:
            year = getYear[-1]
            oldyaer = year
        else:
            year = oldyaer
        if int(year) < SetYear:
            break
        else:
            Attr = (year,u['href'],t.text,)
            MeetPaper.append(Attr)
    # print(MeetPaper)
    return MeetPaper
'''
通过传入关键字进行匹配
# 这里可以做很多有意思的操作，比如我这里是只匹配一个KW
# 通过在这里修改判断，可以做多重关键字的判断
# 比如如果KW1没有匹配到，那能不能匹配kw2 再返回，或者
# KW1匹配到了，再匹配KW2，都匹配到了再返回
'''
def BlockFind(kw, info):
    if re.findall(kw, info, re.I):
        return info

# 这里是爬虫获取会议数据的核心，完成每次会议的论文数据采集
def GetMeetPaperAll(url,KW,SetYear):
    # 第一步：获取到当前期刊的网址主页信息，里面包含有历次的会议信息
    MeetPaper = GetMeetMainInfo(url,SetYear)
    CountMeet = 0  # 这是会议的个数
    CountAllMeetPaper = 0 # 所有会议的论文总数
    CountAllMeetPaperWithKW = 0 #涉及KW的论文总数
    CountMeetWithKW = 0 # 涉及KW的会议总数
    MeetPaperAll = []
    for info in MeetPaper:
        # 数据格式：
        # ('2022', 'https://dblp.uni-trier.de/db/conf/ppopp/ppopp2022.html', "PPoPP '22: 27th ACM SIGPLAN Symposium on Principles and Practice of Parallel Programming, Seoul, Republic of Korea, April 2 - 6, 2022.")
        CountMeet = CountMeet + 1
        year = info[0]
        # 这里是存每次爬的会议信息
        # with open('txt-src/AllMeetInfo.txt','a',encoding='utf-8') as file:
        #     file.write('\n'+'## '+info[-1])
        #     file.write('\n'+'### '+year)
        #     file.close()
        response = GetUrlInfo(info[1])
        soup = BeautifulSoup(response.text, 'lxml')
        temp1 = soup.find_all('li', itemtype="http://schema.org/ScholarlyArticle")
        MeetPaperTitel = []
        CountPaperAll = 0 # 这是当前会议的所有论文的个数
        CountPaperWithKW = 0 # 这是当前会议匹配到关键字的论文总数
        # 第二步，获取历次会议的具体论文信息，包括链接和标题
        for i in temp1:
            u = i.find('div', class_="head").find('a')
            if u == None:
                continue
            t = i.find('span', class_="title", itemprop="name")
            # print(year+', '+u['href']+', '+t.text)
            CountPaperAll = CountPaperAll +1
            MeetPaperTitel.append(t.text)
            if BlockFind(KW,t.text):
                CountPaperWithKW = CountPaperWithKW + 1
                Temp2=(year,u['href'],t.text,)
                # print(Temp2)
                MeetPaperAll.append(Temp2)
        print('会议： '+info[-1]+' 所有论文： '+str(CountPaperAll)+' ,对应KW论文： '+str(CountPaperWithKW))
        CountAllMeetPaper = CountAllMeetPaper + CountPaperAll
        CountAllMeetPaperWithKW = CountAllMeetPaperWithKW + CountPaperWithKW
        if CountPaperWithKW != 0:
            CountMeetWithKW = CountMeetWithKW + 1
        # 这里记录会议的论文信息
        # with open('txt-src/AllMeetInfo.txt','a',encoding='utf-8') as file:
        #     file.write(','+str(CountPaperAll)+'\n')
        #     for i in MeetPaperTitel:
        #         file.write(i+'\n')
        #     file.close()
        # 这里记录每次爬取会议的论文数量，写入日志中
        # with open('txt-src/logs2.txt','a',encoding='utf-8') as f:
        #     f.write('\n'+'## '+info[-1])
        #     f.write('\n' + "### 当前获取论文总数为： ")
        #     f.write(str(CountAllMeetPaper))
        #     f.write(' ,涉及关键字的论文总数为： ')
        #     f.write(str(CountAllMeetPaperWithKW)+'\n')
        #     f.write("### 当前会议总数为： ")
        #     f.write(str(CountMeet))
        #     f.write(' ,涉及关键字的会议总数为： ')
        #     f.write(str(CountMeetWithKW) + '\n')
    print('当前会议总数为： ' +str(CountMeet)+' ,涉及KW的会议总数为： '+str(CountMeetWithKW)+' ,已有论文： '+str(CountAllMeetPaper)+' ,涉及KW 论文： '+str(CountAllMeetPaperWithKW))
    MeetPaperAll.append(CountMeet) #总的会议数
    MeetPaperAll.append(CountMeetWithKW) #涉及KW的会议数
    MeetPaperAll.append(CountAllMeetPaper) #总的论文数
    MeetPaperAll.append(CountAllMeetPaperWithKW) #涉及KW的论文数
    return MeetPaperAll

# 这里的数据url是从表格中获取的
# 设置变量
SetYear = 2006
KW = 'blockchain'

# 这里是数据的来源
wb = load_workbook('中国计算机学会推荐国际学术会议和期刊目录-2019/JournalInfo.xlsx')
sheets = wb.worksheets

if '区块链对应会议' in wb.sheetnames:
    print('区块链对应会议 已存在')
    NewSheets2 = wb['区块链对应会议']
else:
    NewSheets2 = wb.copy_worksheet(wb['会议信息'])
    NewSheets2.title = '区块链对应会议'
wb.save('中国计算机学会推荐国际学术会议和期刊目录-2019/JournalInfo.xlsx')

rows = NewSheets2.rows
columns = NewSheets2.columns
SetRows = 0
# 第一次报错：
# 当前已遍历论文总数为： 176384 ,已遍历对应关键字的论文总数为： 889 ,会议总数为： 2199 ,拥有KW的会议总数为： 217
# 6 ['网络与信息安全', 'A', 'CRYPTO', 'International Cryptology Conference', 'Springer', 'https://dblp.uni-trier.de/db/conf/crypto/']
# 第二次强行停止：
# 当前已遍历论文总数为： 311439 ,已遍历对应关键字的论文总数为： 1435 ,会议总数为： 5629 ,拥有KW的会议总数为： 467
# 6 ['计算机科学理论', 'B', 'CCC', 'IEEE Conference on Computational Complexity', 'IEEE',
''' 
#同样的，当因为网络问题而断开链接后，直接修改参数即可
AllPaperFromMeet = 311439 # 论文总数
AllPaperwithKWFromMeet = 1435 # 对应关键字的论文总数
AllMeet = 5629 # 会议总数
AllMeetWithKW = 467 # 拥有KW的会议总数
'''
AllPaperFromMeet = 0# 论文总数
AllPaperwithKWFromMeet = 0# 对应关键字的论文总数
AllMeet = 0# 会议总数
AllMeetWithKW = 0# 拥有KW的会议总数
for row in rows:
    MainPape = []
    row_val = [col.value for col in row]
    # 对应行的真实列数
    GetCol = GetTrueLen(row_val)
    print(GetCol, row_val)
    print(NewSheets2.cell(row=1,column=GetCol).value)  # 这里输出当前行的最后一位，判断是否已经写入，正常是字符串，然后写完后就是所有论文的数目
    SetRows = SetRows + 1 # 这里这就是行的变化
    if type(NewSheets2.cell(row=SetRows,column=GetCol).value) == int:
        continue
    else:
        # # 先写入CCF对应的会议信息
        # with open('txt-src/AllMeetInfo.txt','a',encoding='utf-8') as file:
        #     file.write('\n' + '# ')
        #     for i in range(GetCol):
        #         if type(row_val[i]) != str:
        #             file.write(str(row_val[i]) + '\t')
        #         else:
        #             file.write(row_val[i] + '\t')
        #     file.close()
        ## 向日志中同样写入会议信息
        # with open('txt-src/logs2.txt', 'a', encoding='utf-8') as file:
        #     file.write('\n' + '# ')
        #     for i in range(GetCol):
        #         if type(row_val[i]) != str:
        #             file.write(str(row_val[i]) + '\t')
        #         else:
        #             file.write(row_val[i] + '\t')
        #     file.close()
        url = NewSheets2.cell(row=SetRows,column=GetCol).value
        MainPape = GetMeetPaperAll(url,KW,SetYear)
        print(MainPape)
        PaperNum = int(MainPape[-1])
        AllMeet = AllMeet + int(MainPape[-4])
        AllMeetWithKW = AllMeetWithKW + int(MainPape[-3])
        AllPaperFromMeet = AllPaperFromMeet + int(MainPape[-2])
        AllPaperwithKWFromMeet = AllPaperwithKWFromMeet + PaperNum
        # 这里写入的格式完全一致
        if PaperNum != 0:
            for j in range(PaperNum):
                NewSheets2.cell(row=SetRows, column=GetCol + j * 2 + 1).value = MainPape[j][0]
                NewSheets2.cell(row=SetRows, column=GetCol + j * 2 + 2).value = MainPape[j][2]
                NewSheets2.cell(row=SetRows, column=GetCol + j * 2 + 2).hyperlink = MainPape[j][1]
        # 不同的是在这里，因为我记录的数据有点多，所以设置的写入多了点
        NewSheets2.cell(row=SetRows, column=GetCol + PaperNum * 2 + 1).value = MainPape[-4]
        NewSheets2.cell(row=SetRows, column=GetCol + PaperNum * 2 + 2).value = MainPape[-3]
        NewSheets2.cell(row=SetRows, column=GetCol + PaperNum * 2 + 3).value = MainPape[-2]
        NewSheets2.cell(row=SetRows, column=GetCol + PaperNum * 2 + 4).value = MainPape[-1]
        wb.save('中国计算机学会推荐国际学术会议和期刊目录-2019/JournalInfo.xlsx')
        print("当前已遍历论文总数为： " + str(AllPaperFromMeet) + ' ,已遍历对应关键字的论文总数为： ' + str(AllPaperwithKWFromMeet) + " ,会议总数为： " + str(
            AllMeet)+ " ,拥有KW的会议总数为： " + str(AllMeetWithKW))

# 当前已遍历论文总数为： 650207 ,已遍历对应关键字的论文总数为： 1605 ,会议总数为： 8458 ,拥有KW的会议总数为： 550

